from django.shortcuts import render, get_object_or_404, redirect
from django.conf import settings
from django.http import HttpResponse
from django.db import IntegrityError
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from .models import CargaAnalisis, MensajesSoporte
from .forms import CustomUserCreationForm, CustomAuthenticationForm, MensajesSoporteForm
import re, os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives

# Create your views here.
def index(request):
    analisis = CargaAnalisis.objects.filter(id_usuario=request.user.id)

    return render(request,'index.html', {
        'analisis': analisis
    })

def signin(request):
    if request.method == 'GET':
        return render(request, 'iniciar_sesion.html',{
            'form': CustomAuthenticationForm
        })
    elif request.method == 'POST':
        user = authenticate(request, username=request.POST['email'], password=request.POST['password'])
        if user is None:
            return render(request, 'iniciar_sesion.html',{
                'form': CustomAuthenticationForm,
                'error': 'El correo o la contraseña es inválido'
            })
        else:
            login(request, user)
            return redirect(to='inicio')

def signup(request):
    if request.method == 'GET':
        return render(request, 'registro.html', {
            'form': CustomUserCreationForm
        })
    elif request.method == 'POST':
        if request.POST['password1'] == request.POST['password2']:
            try:
                user = User.objects.create_user(first_name=request.POST['first_name'], last_name=request.POST['last_name'], email=request.POST['email'], password=request.POST['password1'], username=request.POST['email'])
                user.save()
                login(request, user)
                return redirect(to='inicio')
            except IntegrityError as e:
                print(e)
                return render(request, 'registro.html', {
                    'form': CustomUserCreationForm,
                    'error': 'Hubo un error al intentar guardar la información: ' + str(e)
                })
            
        return render(request, 'registro.html', {
                    'form': CustomUserCreationForm,
                    'error': 'Las contraseñas no coinciden'
                })   

def signout(request):
    logout(request)
    return redirect(to='home')

def home(request):
    return render(request,'layouts/home.html')

def getUser(request, id):
    user = get_object_or_404(User, id = id)
    return HttpResponse("Nombres usuario: %s " % user.nombres)

def createAnalisis(request):
    ruta_archivo = os.path.join(settings.BASE_DIR, 'myapp/static/archivos_txt', 'archivo.txt')
    contenido = ""
    usuario = request.user.id
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        # guardar el archivo en el servidor
        with open(ruta_archivo, 'wb+') as destino:
            for chunk in archivo.chunks():
                destino.write(chunk)

        # leer el contenido del archivo
        with open(ruta_archivo, 'r', encoding='utf-8') as origen:
            contenido = origen.read()

            if request.POST['lenguaje_texto'] == 'en':
                patron1 = r"Generate a citation for this article\s*?\n\s*?PDF:\s*?(.+?)\s*?(?:Me:|$)"
            
                patron2 = r"What is the title of the article\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron3 = r"What is the geographic location (country) of the article\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron4 = r"Give me the textual summary or abstract of the article\s*?\n\s*?PDF:\s*?(.+?)\s*?(?:Me:|$)"

                patron5 = r"What is the topic of the article\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron6 = r"What is the research problem or what will this study solve\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron7 = r"What is the objective of the research? Can be one or many\s*?\n\s*?PDF:\s*?(.+?)\s*?(?:Me:|$)"

                patron8 = r"Give me the theoretical references of the study\s*?\n\s*?PDF:\s*?(.+?)\s*?(?:Me:|$)"

                patron9 = r"What is the type of the methodology\?? It's quantitative, qualitative or mixed\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron10 = r"Taking into account the type of methodology, is it qualitative documentary, quantitative experimental or quantitative correlational\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron11 = r"What is the study methodology and its research paradigm\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron12 = r"What are the results of the study\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron13 = r"What is the conclusion of the study\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron14 = r"What is the APA reference of the article\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"
            else:
                patron1 = r"Genera una cita para este articulo\s*?\n\s*?PDF:\s*?(.+?)\s*?(?:Me:|$)"

                patron2 = r"¿Cual es el titulo del articulo\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron3 = r"¿Cual es la ubicacion geografica \(pais\) del articulo\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron4 = r"Dame el resumen textual o abstracto del articulo\s*?\n\s*?PDF:\s*?(.+?)\s*?(?:Me:|$)"

                patron5 = r"¿Cual es el tema del articulo\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron6 = r"¿Cual es el problema de investigacion o que resolvera este estudio\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron7 = r"¿Cual es el objetivo de la investigacion\?? Puede ser uno o varios\s*?\n\s*?PDF:\s*?(.+?)\s*?(?:Me:|$)"

                patron8 = r"Dame las referencias teoricas del estudio\s*?\n\s*?PDF:\s*?(.+?)\s*?(?:Me:|$)"

                patron9 = r"¿Cual es el tipo de metodologia\?? \¿Es cuantitativa, cualitativa o mixta\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron10 = r"Teniendo en cuenta el tipo de metodologia, \¿es documental cualitativa, experimental cuantitativa o correlacional cuantitativa\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron11 = r"¿Cual es la metodologia del estudio y su paradigma de investigacion\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron12 = r"¿Cuales son los resultados del estudio\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron13 = r"¿Cual es la conclusion del estudio\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

                patron14 = r"¿Cual es la referencia APA del articulo\?.*?PDF:\s*(.+?)\s*?(?:Me:|$)"

            citacion = obtenerRespuesta(patron1, contenido)
            titulo = obtenerRespuesta(patron2, contenido)
            ubicacion = obtenerRespuesta(patron3, contenido)
            resumen = obtenerRespuesta(patron4, contenido)
            tema = obtenerRespuesta(patron5, contenido)
            problema = obtenerRespuesta(patron6, contenido)
            objetivo = obtenerRespuesta(patron7, contenido)
            referentes = obtenerRespuesta(patron8, contenido)
            tipo_metodologia = obtenerRespuesta(patron9, contenido)
            diseno_metodologia = obtenerRespuesta(patron10, contenido)
            metodologia = obtenerRespuesta(patron11, contenido)
            resultados = obtenerRespuesta(patron12, contenido)
            conclusiones =  obtenerRespuesta(patron13, contenido)
            ref_apa = obtenerRespuesta(patron14, contenido)

            i = CargaAnalisis(
                citacion = citacion,
                tema = tema,
                invest_problema = problema,
                invest_objetivo = objetivo, 
                ubicacion = ubicacion,
                titulo_articulo = titulo,
                resumen = resumen,
                problema = problema,
                objetivos = objetivo,
                referentes_teoricos = referentes,
                tipo_metodologia = tipo_metodologia,
                diseno_metodologia = diseno_metodologia,
                metodologia = metodologia,
                resultados = resultados,
                conclusiones = conclusiones,
                lenguaje_texto = request.POST['lenguaje_texto'],
                id_usuario = User.objects.get(id=usuario),
                ref_apa = ref_apa
                )
            try:
                i.save()
                #analisis = CargaAnalisis.objects.all(id_usuario=usuario)
                return redirect(to='herramienta')
            except IntegrityError as e:
                contenido = f"<p>No se pudo guardar la peticion: {str(e)}</p><a href='/'>Volver</a>"
                return HttpResponse(content=contenido,content_type='text/html')
            except Exception as e:
                # Capturar cualquier otra excepción
                contenido = f"<p>Error desconocido. No se pudo guardar: {str(e)}</p><a href='/'>Volver</a>"
                return HttpResponse(content=contenido,content_type='text/html')      

def obtenerRespuesta(patron, contenido):
    match = re.search(patron, contenido, re.DOTALL)

    # Verificar si se encontró una coincidencia y extraer la respuesta
    if match:
        respuesta = match.group(1).strip()
        return respuesta
    else:
        return "No se encontró"
    
def deleteAnalisis(request, id):
    analisis = get_object_or_404(CargaAnalisis, id=id)
    analisis.delete()
    return redirect(to='inicio')

def export_to_excel(request):
    # Crear un libro de trabajo de Excel
    wb = Workbook()

    queryset1 = CargaAnalisis.objects.filter(id_usuario = request.user.id)
    queryset2 = CargaAnalisis.objects.filter(id_usuario = request.user.id)

    ws1 = wb.active
    ws1.title = "Matriz de análisis bibliográfico"

    # Agregar encabezados de columna para el modelo TuModelo
    ws1.append(["No", "Ámbito", "Ubicación", "Idioma","Referencia según APA","Referencia según APA en idioma original","Título del estudio","Tipo","Categoría de estudio","Subcategoría asociada","Resumen del estudio","Problema","Objetivo","Referentes teóricos","Tipo de metodología","Diseño","Metodología","Resultados del estudio","Conclusiones"])

    for index, obj in enumerate(queryset1):
        if obj.lenguaje_texto == 'en':
            idioma = 'Inglés',
        else:
            idioma = 'Español'
        
        ws1.append([index, "", obj.ubicacion, idioma, obj.ref_apa, "", obj.titulo_articulo, "", "", "", obj.resumen, obj.problema, obj.objetivos,
                    str(obj.referentes_teoricos), obj.tipo_metodologia, obj.diseno_metodologia, obj.metodologia, obj.resultados, obj.conclusiones])
        
    for columna in ws1.columns:
        column = columna[0].column_letter
        for celda in columna:
            if celda.row == 1:
                adjusted_width = (len(str(celda.value)) + 2) * 1.2
                ws1.column_dimensions[column].width = adjusted_width
                break    
        
    for columna in ws1.iter_cols():
    # Recorremos todas las filas excepto la primera
        for fila, celdas in enumerate(columna, start=1):
        # Aplicar alineación a la izquierda y permitir salto de línea
            if fila != 1:
                celdas.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)    

    # Crear otro libro para otro conjunto de datos
    ws2 = wb.create_sheet(title="Actividad introduccion")

    # Agregar encabezados de columna para OtroModelo
    ws2.append(["Citation", "Article title", "Topic", "Reseach problem", "Research objetive/s"])

    # Agregar datos de la base de datos de OtroModelo a las filas
    for obj in queryset2:
        ws2.append([obj.citacion, obj.titulo_articulo, obj.tema, obj.problema, obj.objetivos])

    # Guardar el libro de trabajo en un archivo
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Analisis_bibliografico.xlsx'
    wb.save(response)

    return response

def viewUser(request):
    id_usuario = request.user.id
    cantidad_registros = CargaAnalisis.objects.filter(id_usuario=id_usuario).count()
    return render(request, 'perfil.html', {
        'cant_archivos': cantidad_registros
    })

def soporte(request):
    guardado_correctamente = False
    if request.method == 'POST':
        form = MensajesSoporteForm(request.POST, request.FILES)
        if form.is_valid():
            mensaje = form.save(commit=False)
            mensaje.id_usuario = request.user
            mensaje.estado = "Enviado"
            mensaje.save()
            guardado_correctamente = True
            return render(request,'soporte.html', {
                'mensaje': guardado_correctamente
            })
        else:
            return render(request,'soporte.html', {
                'mensaje': guardado_correctamente
            })
    else:
        form = MensajesSoporteForm()
        soportes = MensajesSoporte.objects.filter(id_usuario=request.user.id)

    return render(request, 'soporte.html', {'form': form, 'mensaje': '', 'soportes':soportes})

def recuperar_clave(request):
    if request.method == "POST":
        email = request.POST['email']
        user = User.objects.filter(email=email).first()
        user_exists = User.objects.filter(email=email).exists()
        if user_exists:
            uid = urlsafe_base64_encode(str(user.id).encode('utf-8'))
            token = default_token_generator.make_token(user)
            reset_link = f"https://gestor-bibliografico.onrender.com/restablecer_clave/{uid}/{token}/"
            # Envía el correo electrónico con el enlace de restablecimiento
            subject = 'Solicitud de restablecimiento de contraseña'
            html_content = render_to_string('layouts/restablecer_clave.html', {'reset_link': reset_link})
            text_content = 'Hola,\nAquí está tu enlace de restablecimiento de contraseña: {}'.format(reset_link)
            msg = EmailMultiAlternatives(subject, text_content, None, [email])
            msg.attach_alternative(html_content, "text/html")
            msg.send()
            #messages.success(request, 'Se ha enviado un correo electrónico con instrucciones para restablecer tu contraseña.')
            return redirect('signin')
        else:
            #messages.error(request, 'No hay ninguna cuenta asociada a este correo electrónico.')
            return redirect('recuperar_clave')
    return render(request, 'recuperar_clave.html')

def restablecer_clave(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        # Token válido, permite al usuario restablecer la contraseña
        if request.method == 'POST':
            password = request.POST['password']
            user.set_password(password)
            user.save()
            #messages.success(request, 'Tu contraseña ha sido restablecida con éxito.')
            #return redirect('signin')
            return render(request, 'password_reset_confirm.html')
    else:
        # Token no válido o usuario no encontrado
        #messages.error(request, 'El enlace de restablecimiento de contraseña es inválido o ha caducado.')
        return redirect('signin')

#Ejemplos
def ejemploconcatena(request, nombre):
    return HttpResponse("Hola %s , bienvenido/a" % nombre)

#def users(request):
#    users = list(Users.objects.values())
#    return JsonResponse(users, safe=False)

#def getUser(request, id):
#    user = get_object_or_404(Users, id = id)
#    return HttpResponse("Nombres usuario: %s " % user.nombres)